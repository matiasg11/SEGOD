from __future__ import annotations

import csv
import json
import re
import sys
from datetime import date, datetime
from pathlib import Path

import openpyxl


def clean_name(value: str) -> str:
    value = re.sub(r"[^0-9A-Za-záéíóúÁÉÍÓÚñÑ]+", "_", value).strip("_")
    return value or "hoja"


def serializable(value):
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    return value


def find_header(sheet) -> int | None:
    if sheet.title.startswith(("LAB-", "DC-")):
        return 5 if any(sheet.cell(5, col).value not in (None, "") for col in range(1, sheet.max_column + 1)) else None
    if any(sheet.cell(1, col).value not in (None, "") for col in range(1, sheet.max_column + 1)):
        return 1
    candidates = []
    for row_number in range(1, min(sheet.max_row, 20) + 1):
        values = [sheet.cell(row_number, col).value for col in range(1, min(sheet.max_column, 40) + 1)]
        count = sum(value not in (None, "") for value in values)
        if count >= 2:
            candidates.append((count, row_number))
    if not candidates:
        return None
    return max(candidates, key=lambda item: (item[0], item[1]))[1]


def main() -> None:
    if len(sys.argv) != 3:
        raise SystemExit("Uso: python extract_appsheet_excel.py origen.xlsx carpeta_salida")
    source = Path(sys.argv[1])
    output = Path(sys.argv[2])
    output.mkdir(parents=True, exist_ok=True)
    workbook = openpyxl.load_workbook(source, data_only=True)
    manifest = {"source": str(source), "extracted_at": datetime.now().isoformat(), "sheets": []}

    for sheet in workbook.worksheets:
        header_row = find_header(sheet)
        if header_row is None:
            continue
        headers = [sheet.cell(header_row, col).value for col in range(1, sheet.max_column + 1)]
        last_col = max((i for i, value in enumerate(headers, 1) if value not in (None, "")), default=0)
        headers = [str(value).strip() if value not in (None, "") else f"columna_{i}" for i, value in enumerate(headers[:last_col], 1)]
        rows = []
        for row_number in range(header_row + 1, sheet.max_row + 1):
            values = [serializable(sheet.cell(row_number, col).value) for col in range(1, last_col + 1)]
            if any(value not in (None, "") for value in values):
                rows.append(values)
        destination = output / f"{clean_name(sheet.title)}.csv"
        with destination.open("w", encoding="utf-8-sig", newline="") as handle:
            writer = csv.writer(handle)
            writer.writerow(headers)
            writer.writerows(rows)
        manifest["sheets"].append({"sheet": sheet.title, "header_row": header_row, "rows": len(rows), "file": destination.name})

    (output / "manifest.json").write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(manifest, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()


