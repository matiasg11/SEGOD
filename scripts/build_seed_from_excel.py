from __future__ import annotations

import sys
from pathlib import Path

import openpyxl


def sql(value):
    if value is None or value == "":
        return "null"
    if isinstance(value, bool):
        return "true" if value else "false"
    if hasattr(value, "isoformat"):
        return "'" + value.isoformat() + "'"
    return "'" + str(value).replace("'", "''") + "'"


def yes(value):
    return str(value or "").strip().lower() in {"sí", "si", "s", "yes", "true"}


def raw_key(name):
    lowered = str(name).lower()
    mapping = {
        "marcado": "marcado", "talle": "medicion", "medidas": "medicion",
        "desteridad": "desteridad", "cromo": "cromo_vi", "ph": "ph",
        "impacto": "impacto", "corte": "corte_cuchilla", "perfor": "perforacion",
        "rasgado": "rasgado", "abrasi": "abrasion",
    }
    return next((value for key, value in mapping.items() if key in lowered), None)


def main():
    source = Path(sys.argv[1])
    destination = Path(sys.argv[2])
    wb = openpyxl.load_workbook(source, data_only=True)
    statements = ["-- Generado desde Certificación de Laboratorio (1).xlsx", "begin;"]

    staff = wb.worksheets[5]
    for row in staff.iter_rows(min_row=6, values_only=True):
        if not row[0] or not row[1]:
            continue
        values = [row[0], row[1], row[2], row[3], yes(row[5]), yes(row[7]), yes(row[9]), row[6], row[8], yes(row[10]), row[11], row[12]]
        statements.append("insert into public.staff (legacy_id,full_name,job_title,area,can_run_tests,can_approve_reports,can_manage_records,authorized_tests,competence_evidence,impartiality_declared,status,notes) values (" + ",".join(sql(v) for v in values) + ") on conflict (legacy_id) do update set full_name=excluded.full_name,job_title=excluded.job_title,area=excluded.area,can_run_tests=excluded.can_run_tests,can_approve_reports=excluded.can_approve_reports,can_manage_records=excluded.can_manage_records,authorized_tests=excluded.authorized_tests,competence_evidence=excluded.competence_evidence,impartiality_declared=excluded.impartiality_declared,status=excluded.status,notes=excluded.notes;")

    tests = wb.worksheets[0]
    for row in tests.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        values = [row[0], row[1], row[2], row[3], yes(row[4]), raw_key(row[0])]
        statements.append("insert into public.test_catalog (name,standard,method,required_equipment,available_in_house,raw_schema_key) values (" + ",".join(sql(v) for v in values) + ") on conflict (name,standard) do update set method=excluded.method,required_equipment=excluded.required_equipment,available_in_house=excluded.available_in_house,raw_schema_key=excluded.raw_schema_key;")

    legacy_equipment = {}
    for row in wb.worksheets[1].iter_rows(min_row=2, values_only=True):
        if row[0]:
            legacy_equipment[str(row[0])] = row
    equipment = wb.worksheets[8]
    for row in equipment.iter_rows(min_row=6, values_only=True):
        if not row[0] or not row[1]:
            continue
        old = legacy_equipment.get(str(row[0]), (None,) * 11)
        values = [row[0], row[1], row[2], row[3], row[4], yes(row[5]), yes(row[6]), row[7], row[8], row[9], row[10], old[6], old[7], old[3], old[4], old[5], old[8], old[10], row[11]]
        statements.append("insert into public.equipment (code,name,category,main_use,controlled_magnitude,requires_calibration,requires_verification,frequency,status,location,responsible_name,brand,model,acquired_at,last_calibrated_at,last_verified_at,expires_at,certificate_url,notes) values (" + ",".join(sql(v) for v in values) + ") on conflict (code) do update set name=excluded.name,category=excluded.category,main_use=excluded.main_use,controlled_magnitude=excluded.controlled_magnitude,requires_calibration=excluded.requires_calibration,requires_verification=excluded.requires_verification,frequency=excluded.frequency,status=excluded.status,location=excluded.location,responsible_name=excluded.responsible_name,brand=excluded.brand,model=excluded.model,acquired_at=excluded.acquired_at,last_calibrated_at=excluded.last_calibrated_at,last_verified_at=excluded.last_verified_at,expires_at=excluded.expires_at,certificate_url=excluded.certificate_url,notes=excluded.notes;")

    statements.extend(["commit;", ""])
    destination.parent.mkdir(parents=True, exist_ok=True)
    destination.write_text("\n".join(statements), encoding="utf-8")
    print(f"Seed creado: {destination} ({len(statements) - 3} registros)")


if __name__ == "__main__":
    main()
